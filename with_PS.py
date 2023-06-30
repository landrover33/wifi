from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl import Workbook
import subprocess
import time
import os


def get_current_dir():
    current_file_path = os.path.realpath(__file__)
    current_directory = os.path.dirname(current_file_path)
    return current_directory


# print("Current Directory:", get_current_dir())


# ================================================================


def append_data_to_excel(file_path, date_value, signal_value):
    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        workbook = Workbook()

    sheet = workbook.active
    last_row = sheet.max_row + 1

    sheet[f"A{last_row}"] = date_value
    sheet[f"B{last_row}"] = signal_value

    workbook.save(file_path)

# ================================================================


script_path = get_current_dir() + os.sep + "wifi_signal.ps1"


def get_wifi_signal(script_path):
    # Execute the PowerShell script
    result = subprocess.run(
        ['powershell.exe', '-File', script_path], capture_output=True, text=True)

    # Access the output and error, if any
    output = result.stdout.strip()
    error = result.stderr.strip()

    # Display the output and error
    if output:
        # print("Output:", output)
        return output
    if error:
        # print("Error:", error)
        return error


# =============================================
for i in range(10):
    time.sleep(1)
    signal_value = get_wifi_signal(script_path)
    append_data_to_excel("signal.xls", "2023-06-25", signal_value)
    print("Output:", signal_value)
